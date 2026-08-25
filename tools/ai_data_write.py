"""
Writes the JSON aggregate report from tools/ai_data_report.js into AI.DATA.xlsx's existing sheets
(CONJOB, ABCM) -- fills in cells by matching existing row/column headers, never touches the sheet
layout/formatting itself.

CONJOB sheet: two pairs of 8x9 tables sharing the same layout (CON001A/002A/003A/004A/001B/002B/003B/
004B rows x JOB001..JOB008 columns) -- "試行回数" (trial count), "平均得点" (average score, QST VP
excluded -- see report['conjob'][].avgScore's own doc in ai_data_report.js), (2026-08-09, per user
request) "QST平均得点" (QST's own average VP contribution, same row/col layout, a separate table further
down the sheet -- report['conjob'][].avgQstScore), and (2026-08-20, per user request) "平均順位" (average
final placement 1-4, same row/col layout again -- report['conjob'][].avgRank). The JOB column headers have no tier suffix (JOB cards
only ever have an 'A' face -- see game-state.js's splitCardId), so a jobFaceId like "JOB005A" is matched
by stripping the trailing letter. A fifth, single-row table (2026-08-07, per user spec) lives at row 30
("使用回数" in column A, JOB001..JOB008 in B:I, the user's own pre-existing layout -- found by
inspection, same as the tables above): one aggregate "usage" value per JOB (not per-round) -- see
report['job']'s own meaning, set by tools/ai_data_report.js's jobEntry doc.

ABCM sheet: one table, ID column (every A/B/C tier-A and tier-B face, M001-M012, plus "D" for "gained a
new colored die") x 試行回数1R-4R / 平均得点1R-4R / QST平均得点1R-4R / 使用回数1-4 / 平均順位 / 平均順位1R-4R
columns (2026-08-20: 平均順位 is a SINGLE column -- the overall average final placement across every round
that card was built in, see report['abcm'][id].avgRank -- and 平均順位1R-4R are the per-round breakdown
added right alongside it, see report['abcm'][id][round].avgRank; both are written). Column
positions are looked up by their own header text (find_col, below), not hardcoded -- the exact same
"never desync from the sheet's actual current layout" reasoning as CONJOB's find_table/find_header_row
(2026-08-09: this sheet's columns were reshuffled to interleave the new QST平均得点{r}R columns between
each existing 平均得点{r}R one, which would have silently written into the wrong columns had the old
hardcoded indices been kept). The 使用回数 columns (2026-08-07, per user spec, also the user's own
pre-existing layout addition) only ever get a value for rows report['abcm'][id][round]['avgUsage'] is
non-null for (B301A/B301B, renamed from B008A/B008B by the 2026-08-24 SHOP201-203 rework's card
renumbering, and the 8 A-deck fee-generating cards -- see tools/ai_data_report.js's isUsageEligible;
B301A means "LVUP success rate", B301B means "max(天,地,人) emblem count at GAME_END" (2026-08-17,
following B301B's own PASSIVE metric change), per user spec 2026-08-12) --
every other row's cells are cleared to blank, not written as 0, so a card with no "usage" concept
at all never shows a misleading number.

Usage: python tools/ai_data_write.py [jsonPath] [xlsxPath]

Row/column headers are matched by NAME, not by ID (2026-08-18, per the user's own edit to AI.DATA.xlsx:
"NAMEや能力など直しました" -- both CONJOB's row labels (was "CON001A" etc, now "祝福" etc) and column
labels (was "JOB001" etc, now "社交家" etc), plus ABCM's ID column (was "A001A" etc, now "小麦畑の支配"
etc, except the literal "D" row, which isn't a real card and has no NAME to look up), switched from
physical/face IDs to the card's actual Japanese name. report['conjob']/['job']/['abcm'] still key by ID
(that's what the AI engine/data-loader deal in), so every ID this script looks up against the sheet is
first translated to its NAME via data/game.json (loaded fresh here, independent of the JSON report) --
see NAME_BY_* below. A missing lookup falls back to the raw ID unchanged (same effect as the old
behavior: it just won't match anything in the NAME-keyed sheet either, so still ends up skipped/reported
rather than silently mis-filed).
"""

import json
import re
import sys
import openpyxl
from openpyxl.utils import get_column_letter

JSON_PATH = sys.argv[1] if len(sys.argv) > 1 else 'output/ai_data_report.json'
XLSX_PATH = sys.argv[2] if len(sys.argv) > 2 else 'AI.DATA.xlsx'

with open(JSON_PATH, 'r', encoding='utf-8') as f:
    report = json.load(f)

with open('data/game.json', 'r', encoding='utf-8') as f:
    game_data = json.load(f)
NAME_BY_CON_FACE = {row['ID']: row['NAME'] for row in game_data['CON']}
NAME_BY_JOB = {row['ID']: row['NAME'] for row in game_data['JOB']}
NAME_BY_CARD_FACE = {row['ID']: row['NAME'] for sheet in ('A', 'B', 'C', 'M') for row in game_data[sheet]}

wb = openpyxl.load_workbook(XLSX_PATH)

# ---------------------------------------------------------------------------
# CONJOB sheet
# ---------------------------------------------------------------------------
ws = wb['CONJOB']

def find_table(header_row):
    """Given the row index of a table's own header row, returns {con_face_id: row_idx} and
    {job_physical_id: col_idx} (job ids stripped to e.g. "JOB005", matching the sheet's own headers)."""
    row_map = {}
    r = header_row + 1
    while ws.cell(row=r, column=1).value:
        row_map[ws.cell(row=r, column=1).value] = r
        r += 1
    col_map = {}
    c = 2
    while ws.cell(row=header_row, column=c).value:
        col_map[ws.cell(row=header_row, column=c).value] = c
        c += 1
    return row_map, col_map

def find_header_row(label):
    """Finds label ("試行回数" or "平均得点") in column A -- searched fresh each run rather than a
    hardcoded row number, so inserting/removing CON rows (e.g. 2026-08-03's CON005A/B addition) never
    silently desyncs this script from the sheet's actual current layout. Scans the whole used range
    (not stopping at the first blank row) since the two tables are separated by blank spacer rows."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    raise ValueError(f'Could not find a "{label}" header in column A of the CONJOB sheet')

count_rows, count_cols = find_table(find_header_row('試行回数'))
avg_header_row = find_header_row('平均得点')
avg_rows, avg_cols = find_table(avg_header_row)
# QST平均得点 (2026-08-09, per user request) -- a third table, same row/col layout as the two above,
# further down the sheet (the user's own pre-existing addition, found the same dynamic way).
qst_avg_rows, qst_avg_cols = find_table(find_header_row('QST平均得点'))
# 平均順位 (2026-08-20, per user request: "これを参考に平均順位を書き出すようにしてほしい") -- a fourth
# table, same CON row / JOB column layout as the three above (no extra K/N-style columns, unlike 平均得点's
# own VPペナルティ平均 -- confirmed by inspection), also the user's own pre-existing addition.
rank_avg_rows, rank_avg_cols = find_table(find_header_row('平均順位'))

def find_col_in_row(row, label):
    """Same idea as find_col (below, ABCM sheet) but for a given CONJOB header row rather than always
    row 1 -- used for the 平均得点 table's own extra columns (K/N) that live past its JOB008 column."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=c).value == label:
            return c
    raise ValueError(f'Could not find column "{label}" in CONJOB row {row}')

# Clears only the JOB001..JOB008 data cells this script actually writes (2026-08-03, per user feedback:
# "すでにある数字を上書きして大丈夫です" for a fresh N-game run) -- without this, a combination that
# simply didn't occur in *this* run would silently keep whatever number an earlier (possibly pre-bugfix)
# run left behind, mixing two different AI behaviors/seed counts in the same sheet.
# Bounded to max(count_cols.values()) (2026-08-07, fixing a bug reported by the user: "CONJOBシートの
# セルL16からM25までもコピペされるようにしてください") -- this used to sweep all the way to
# ws.max_column, which also wiped columns L/M (rows 16-25, alongside the 平均得点 table): the user's own
# hand-maintained reference notes (each CON face's own ONCE-effect DSL + INST description text, e.g. row
# 16 = CON001A's "ADD(6K)" / "資源◯の上限7個..."), unrelated to anything this script computes or writes.
last_job_col = max(count_cols.values())
for row_map in (count_rows, avg_rows, qst_avg_rows, rank_avg_rows):
    for r in row_map.values():
        for c in range(2, last_job_col + 1):
            ws.cell(row=r, column=c).value = None

written = 0
skipped = []
for entry in report['conjob']:
    con_face_id = entry['con']
    job_physical_id = re.sub(r'[A-Z]$', '', entry['job'])  # "JOB005A" -> "JOB005"
    con_name = NAME_BY_CON_FACE.get(con_face_id, con_face_id)
    job_name = NAME_BY_JOB.get(job_physical_id, job_physical_id)
    if con_name not in count_rows or job_name not in count_cols:
        skipped.append((con_face_id, entry['job']))
        continue
    ws.cell(row=count_rows[con_name], column=count_cols[job_name], value=entry['count'])
    ws.cell(row=avg_rows[con_name], column=avg_cols[job_name], value=round(entry['avgScore'], 2))
    ws.cell(row=qst_avg_rows[con_name], column=qst_avg_cols[job_name], value=round(entry['avgQstScore'], 2))
    ws.cell(row=rank_avg_rows[con_name], column=rank_avg_cols[job_name], value=round(entry['avgRank'], 2))
    written += 1

print(f'CONJOB: wrote {written} combinations, skipped {len(skipped)}: {skipped}')

# K: per-CON marginal average across the 8 JOB columns (2026-08-07, per user request: "JOBCONシートの
# J、K列も出力できるようにしてください"; J column abolished and K's formula corrected 2026-08-10, per
# user report: "試行回数が0のものがあると平均値が著しくずれてしまいます" -- the original K=J/8 divided
# by a FIXED 8 regardless of how many of those 8 JOB columns actually had any games this run. A blank
# (0-trial) 平均得点 cell contributes nothing to J's SUM, but still counted toward the /8 divisor, so
# every 0-trial JOB column dragged K down even though it has no data at all. Excel's own AVERAGE()
# already ignores blank cells when computing its average (only counts actual numeric cells), which is
# exactly "the same marginal average, but over however many JOB columns actually have data" -- so K now
# calls AVERAGE(...) directly on the B:I row instead of dividing a separately-computed SUM, and J (which
# only ever existed to feed that division) is no longer written. Still a live Excel formula, not a
# pre-computed Python value, so it keeps recalculating whenever this script rewrites the underlying B:I
# cells on a later run. Written for every CON row unconditionally (not just ones report['conjob'] had
# data for this run) and for the QST平均得点 table's own rows too (2026-08-09) -- same per-CON-row
# marginal, just at that table's own row positions.
j_col = last_job_col + 1
k_col = last_job_col + 2
first_col_letter = get_column_letter(min(avg_cols.values()))
last_col_letter = get_column_letter(max(avg_cols.values()))
for row_map in (avg_rows, qst_avg_rows):
    for con_face_id, r in row_map.items():
        ws.cell(row=r, column=j_col, value=None)  # abolished -- see this block's own comment
        # IFERROR-wrapped (found while verifying this change): AVERAGE() of an all-blank row (a CON with
        # zero games across all 8 JOBs this run) is #DIV/0!, not 0 -- surfacing a spreadsheet full of
        # error cells for every untested CON would be its own kind of misleading "wrong average" this
        # fix is meant to get away from, so those rows blank out instead.
        ws.cell(row=r, column=k_col, value=f'=IFERROR(AVERAGE({first_col_letter}{r}:{last_col_letter}{r}),"")')

print(f'CONJOB K: wrote per-CON marginal average formulas for {len(avg_rows)} + {len(qst_avg_rows)} rows (J column abolished)')

# N (平均得点 table only, 2026-08-15, per user request: "VPペナルティがあるものは　その平均も出るように
# してください") -- one value per CON face, independent of JOB, so it doesn't fit the (CON,JOB) grid
# these tables are otherwise shaped around; found by its own header text like every other column here
# rather than a hardcoded position (see this file's own top-of-file doc on why). Only ever written for
# the CON faces report['conVpPenalty'] actually has data for (driven entirely by whatever
# tools/ai_data_report.js's CON_VP_PENALTY_FACES contains -- never hardcoded here, so this list is not
# duplicated/kept in sync here) -- every other row's cell is cleared to blank, same "no misleading number
# for a card with no such concept" policy the 使用回数 columns already use.
vp_penalty_col = find_col_in_row(avg_header_row, 'VPペナルティ平均')
for r in avg_rows.values():
    ws.cell(row=r, column=vp_penalty_col).value = None

vp_penalty_written = 0
vp_penalty_skipped = []
for con_face_id, entry in report.get('conVpPenalty', {}).items():
    con_name = NAME_BY_CON_FACE.get(con_face_id, con_face_id)
    if con_name not in avg_rows:
        vp_penalty_skipped.append(con_face_id)
        continue
    if entry['avgVpPenalty'] is not None:
        ws.cell(row=avg_rows[con_name], column=vp_penalty_col, value=round(entry['avgVpPenalty'], 2))
    vp_penalty_written += 1

print(f'CONJOB N (VPペナルティ平均): wrote {vp_penalty_written} CON faces, skipped {len(vp_penalty_skipped)}: {vp_penalty_skipped}')

# Third table: a single "使用回数" row (2026-08-07, per user spec, at the user's own pre-existing row --
# see this file's own top-of-file doc) reusing the same JOB001..JOB008 column positions as the two tables
# above (count_cols) rather than re-deriving them, since there's no separate header row of its own.
job_usage_row = find_header_row('使用回数')
for c in range(2, last_job_col + 1):
    ws.cell(row=job_usage_row, column=c).value = None

job_written = 0
job_skipped = []
for job_face_id, entry in report.get('job', {}).items():
    job_physical_id = re.sub(r'[A-Z]$', '', job_face_id)
    job_name = NAME_BY_JOB.get(job_physical_id, job_physical_id)
    if job_name not in count_cols:
        job_skipped.append(job_face_id)
        continue
    if entry['avgUsage'] is not None:
        ws.cell(row=job_usage_row, column=count_cols[job_name], value=round(entry['avgUsage'], 2))
    job_written += 1

print(f'CONJOB row {job_usage_row} (JOB使用回数): wrote {job_written} JOBs, skipped {len(job_skipped)}: {job_skipped}')

# Per-JOB TRUE weighted average score / QST平均得点 / 平均順位 (2026-08-20, per user request: "この計算式
# で出る平均は実際の平均とずれると思います...実際の平均にしてもらうことはできますか") -- replaces the
# sheet's own unweighted "=AVERAGE(12 already-averaged per-CON cells)" formula at these rows with the real
# per-JOB weighted value computed in ai_data_report.js's jobEntry (see its own doc for why the unweighted
# version can be a real, not-always-negligible bias). Each row sits 2 rows below its own table's last CON
# row (a blank spacer row, then the marginal row) -- the same spacing the user's own sheet already uses
# for every table here, so found the same dynamic way rather than hardcoded row numbers. Row 32 (JOB's own
# QST平均得点, ="=B51" etc) and row 33 (JOB's own combined total, "=B31+B32") are left as formulas --
# writing a correct value into row 51 (and 31) automatically flows through both, no need to touch them.
job_score_row = max(avg_rows.values()) + 2
job_qst_score_row = max(qst_avg_rows.values()) + 2
job_rank_row = max(rank_avg_rows.values()) + 2
for r in (job_score_row, job_qst_score_row, job_rank_row):
    for c in range(2, last_job_col + 1):
        ws.cell(row=r, column=c).value = None

job_avg_written = 0
job_avg_skipped = []
for job_face_id, entry in report.get('job', {}).items():
    job_physical_id = re.sub(r'[A-Z]$', '', job_face_id)
    job_name = NAME_BY_JOB.get(job_physical_id, job_physical_id)
    if job_name not in count_cols:
        job_avg_skipped.append(job_face_id)
        continue
    col = count_cols[job_name]
    if entry.get('avgScore') is not None:
        ws.cell(row=job_score_row, column=col, value=round(entry['avgScore'], 2))
    if entry.get('avgQstScore') is not None:
        ws.cell(row=job_qst_score_row, column=col, value=round(entry['avgQstScore'], 2))
    if entry.get('avgRank') is not None:
        ws.cell(row=job_rank_row, column=col, value=round(entry['avgRank'], 2))
    job_avg_written += 1

print(f'CONJOB rows {job_score_row}/{job_qst_score_row}/{job_rank_row} (JOBの真の加重平均: 平均得点/QST平均得点/平均順位): wrote {job_avg_written} JOBs, skipped {len(job_avg_skipped)}: {job_avg_skipped}')

# ---------------------------------------------------------------------------
# ABCM sheet
# ---------------------------------------------------------------------------
ws2 = wb['ABCM']
id_rows = {}
r = 2
while ws2.cell(row=r, column=1).value:
    id_rows[ws2.cell(row=r, column=1).value] = r
    r += 1

# Column positions looked up by their own row-1 header text, not hardcoded (2026-08-09 -- see this
# file's own top-of-file doc for why: this sheet's columns were reshuffled once already to interleave
# the new QST平均得点{r}R columns, which would have silently desynced a hardcoded layout).
def find_col(label):
    for c in range(1, ws2.max_column + 1):
        if ws2.cell(row=1, column=c).value == label:
            return c
    raise ValueError(f'Could not find column "{label}" in row 1 of the ABCM sheet')

COUNT_COLS = {r: find_col(f'試行回数{r}R') for r in (1, 2, 3, 4)}
AVG_COLS = {r: find_col(f'平均得点{r}R') for r in (1, 2, 3, 4)}
QST_AVG_COLS = {r: find_col(f'QST平均得点{r}R') for r in (1, 2, 3, 4)}
USAGE_COLS = {r: find_col(f'使用回数{r}') for r in (1, 2, 3, 4)}
# 平均順位 (2026-08-20, per user request) -- the overall average rank across every round that card was
# ever built in, kept as a single column (not broken out per round) alongside...
RANK_AVG_COL = find_col('平均順位')
# ...the per-round 平均順位1R-4R columns the user added right after it (2026-08-20, per user request:
# "AI.DATA平均順位1R-4Rを追加しました出せるようにおねがい") -- both forms are written; see
# tools/ai_data_report.js's own doc for why abcmOut[id][round].avgRank and abcmOut[id].avgRank coexist.
RANK_AVG_COLS = {r: find_col(f'平均順位{r}R') for r in (1, 2, 3, 4)}

# Same clear-before-write as the CONJOB sheet above, same reason.
for r in id_rows.values():
    for c in list(COUNT_COLS.values()) + list(AVG_COLS.values()) + list(QST_AVG_COLS.values()) + list(USAGE_COLS.values()) + [RANK_AVG_COL] + list(RANK_AVG_COLS.values()):
        ws2.cell(row=r, column=c).value = None

abcm_written = 0
abcm_skipped = []
for card_id, entry in report['abcm'].items():
    card_name = NAME_BY_CARD_FACE.get(card_id, card_id)  # 'D' has no card NAME, passes through unchanged
    if card_name not in id_rows:
        abcm_skipped.append(card_id)
        continue
    row = id_rows[card_name]
    for round_str in ('1', '2', '3', '4'):  # skip the sibling 'avgRank' key -- handled separately below
        cell = entry[round_str]
        round_num = int(round_str)
        ws2.cell(row=row, column=COUNT_COLS[round_num], value=cell['count'])
        if cell['avgScore'] is not None:
            ws2.cell(row=row, column=AVG_COLS[round_num], value=round(cell['avgScore'], 2))
        if cell.get('avgQstScore') is not None:
            ws2.cell(row=row, column=QST_AVG_COLS[round_num], value=round(cell['avgQstScore'], 2))
        if cell.get('avgUsage') is not None:
            ws2.cell(row=row, column=USAGE_COLS[round_num], value=round(cell['avgUsage'], 2))
        if cell.get('avgRank') is not None:
            ws2.cell(row=row, column=RANK_AVG_COLS[round_num], value=round(cell['avgRank'], 2))
    if entry.get('avgRank') is not None:
        ws2.cell(row=row, column=RANK_AVG_COL, value=round(entry['avgRank'], 2))
    abcm_written += 1

print(f'ABCM: wrote {abcm_written} card rows, skipped {len(abcm_skipped)}: {abcm_skipped}')

# ---------------------------------------------------------------------------
# HighScores sheet (2026-08-04, per user feedback: "20点以上の点数があった時 その得点を取ったAIが何を
# したのか確認できるように...ログは一つのエクセルファイルにまとめる") -- one row per player (all 4, not
# just whoever crossed the threshold, per "4人とも記録 それぞれの得点も") for every game where at least
# one player's score reached report['highScoreThreshold']. Deliberately limited to score-relevant fields
# only (CON/JOB/initial RESOURCE/builds by round, per "スコアに直結する行動だけに絞ります") -- not a full
# move-by-move log.
# ---------------------------------------------------------------------------
if 'HighScores' in wb.sheetnames:
    del wb['HighScores']
ws3 = wb.create_sheet('HighScores')
ws3.append(['Seed', 'Player', 'Score', 'CON', 'JOB', 'Resources', 'R1 Builds', 'R2 Builds', 'R3 Builds', 'R4 Builds'])
for row in report.get('highScoreRows', []):
    ws3.append([
        row['seed'],
        row['playerId'],
        row['score'],
        row['con'],
        row['job'],
        ','.join(row['resources']),
        ','.join(row['builds1']),
        ','.join(row['builds2']),
        ','.join(row['builds3']),
        ','.join(row['builds4']),
    ])
print(f"HighScores: wrote {len(report.get('highScoreRows', []))} player-rows (threshold={report.get('highScoreThreshold')})")

wb.save(XLSX_PATH)
print(f'Saved {XLSX_PATH} (based on {report["gamesRun"]} games)')
