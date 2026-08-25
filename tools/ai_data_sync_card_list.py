"""
One-time-per-addition structural sync for AI.DATA.xlsx, run by hand whenever data/game.json's card/JOB
list has drifted from what the ABCM/CONJOB sheets already know about (2026-08-25, per user request:
"AIDATAででる情報を今のカードリストに合わせてほしい"). Like ai_data_add_missing_job_columns.py, this
changes sheet STRUCTURE (new rows, renamed headers) rather than filling in report values -- run this
first, then the normal ai_data_report.js + ai_data_write.py pipeline to populate the new rows/columns.

ABCM sheet: appends one row (in column A, matching ai_data_write.py's own NAME-based row lookup) for
every current A/B/C/M card face whose NAME isn't already a row label -- always appended after the last
existing row (contiguous, no gaps, since both this script's and ai_data_write.py's row scan is a plain
`while cell(row, 1).value` loop from row 2). Never touches or reorders existing rows.

CONJOB sheet: renames JOB column headers via RENAMES below when a JOB's NAME changed in game.json but
the sheet still shows the old text (found here: JOB005 "酒場の主人" -> "権力者"). Column position is
unchanged, just the header cell text -- safe since ai_data_write.py looks up columns by header text.

Usage: python tools/ai_data_sync_card_list.py [xlsxPath]
"""

import sys
import json
import openpyxl

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else 'AI.DATA.xlsx'
GAME_JSON_PATH = 'data/game.json'

# Old JOB column header text -> current game.json NAME, applied to every occurrence in row 1 of the
# CONJOB sheet's JOB header rows (试行回数/平均得点/QST平均得点/平均順位/使用回数 all share the same
# JOB001..JOB011 header text). Add an entry here whenever a JOB/CON's NAME changes in game.xlsx and this
# script reports it as still-missing after a re-run.
RENAMES = {
    '酒場の主人': '権力者',  # JOB005, renamed in game.xlsx before 2026-08-25
}

with open(GAME_JSON_PATH, 'r', encoding='utf-8') as f:
    game_data = json.load(f)

wb = openpyxl.load_workbook(XLSX_PATH)

# ---------------------------------------------------------------------------
# CONJOB: rename stale JOB column headers
# ---------------------------------------------------------------------------
ws = wb['CONJOB']
renamed = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value in RENAMES:
            cell.value = RENAMES[cell.value]
            renamed += 1
print(f'CONJOB: renamed {renamed} header cell(s)')

# ---------------------------------------------------------------------------
# ABCM: append rows for any current A/B/C/M card face not yet in the sheet
# ---------------------------------------------------------------------------
ws2 = wb['ABCM']
existing_names = set()
r = 2
while ws2.cell(row=r, column=1).value:
    existing_names.add(ws2.cell(row=r, column=1).value)
    r += 1
next_row = r

added = []
for sheet in ('A', 'B', 'C', 'M'):
    for card in game_data[sheet]:
        name = card['NAME']
        if name not in existing_names:
            ws2.cell(row=next_row, column=1, value=name)
            existing_names.add(name)
            added.append((card['ID'], name))
            next_row += 1

print(f'ABCM: appended {len(added)} row(s): {added}')

wb.save(XLSX_PATH)
print(f'Saved {XLSX_PATH}.')
print('Run a fresh AI battle (node tools/ai_data_report.js <N> && python tools/ai_data_write.py) to fill in the new rows/columns\' data.')
