"""
One-time-per-addition structural migration for AI.DATA.xlsx's CONJOB sheet (2026-08-18, per user report:
"run_ai_battleをしたときのAIDATAが以前の表記になっているため今のデータに沿って直してください" --
JOB009/010/011 had no header column at all in the CONJOB sheet, so ai_data_write.py's own "match existing
headers, never touch layout" design silently skipped every combination involving them).

Unlike ai_data_write.py (which only ever fills values into an already-correct layout, run after every AI
battle), this script changes the sheet's actual STRUCTURE -- inserting new JOB columns -- so it's meant to
be run by hand, once, whenever a new JOB face is added to game.json and AI.DATA.xlsx hasn't caught up yet
(not part of the regular ai_data_report.js/ai_data_write.py pipeline). Always inserts new columns
immediately after the current last JOB column (keeping every JOB column contiguous, which
ai_data_write.py's own find_table/find_col column-scan requires), then extends the two hand-maintained
per-JOB aggregate row-triples (SUM / SUM÷12 average / short TAP-ability label -- rows 30-32 for the
平均得点 table, 49-51 for QST平均得点) to cover the new columns too.

What this does NOT do: fix the per-CON marginal-average column (the old "K" column, now shifted right by
however many columns were inserted) -- its AVERAGE(...) formula still only covers the *old* JOB range
until the next real `python tools/ai_data_report.js ... && python tools/ai_data_write.py` run overwrites
it with the correct, now-wider range (ai_data_write.py already recomputes that range dynamically every
run, so this is harmless, just cosmetically stale until then).

Usage: python tools/ai_data_add_missing_job_columns.py [xlsxPath]
"""

import sys
import json
import re
import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else 'AI.DATA.xlsx'
GAME_JSON_PATH = 'data/game.json'

# Short TAP-ability labels for the aggregate block's own label row (rows 32/51), matching the existing
# style (e.g. JOB001's "TAP　Z", JOB002's "建築→K") -- per user request ("自分で短い説明を提案してほしい").
# Only ever used for a JOB id this script is actually about to insert a new column for; a JOB added later
# without an entry here just gets a blank label cell instead of erroring.
PROPOSED_LABELS = {
    'JOB009': '空AREA→ABC',   # 開拓者: places a color die on an empty AREA -> random A/B/C
    'JOB010': '2K→LVUP',      # 革命家: pay 2K, upgrade a card
    'JOB011': 'LVAREA→K/VP',  # 地主: places on an upgraded AREA -> K, or VP if already there
}

with open(GAME_JSON_PATH, 'r', encoding='utf-8') as f:
    game_data = json.load(f)
current_job_ids = [row['ID'] for row in game_data['JOB']]

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['CONJOB']


def find_header_row(label):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    raise ValueError(f'Could not find a "{label}" header in column A of the CONJOB sheet')


def job_header_cols(header_row):
    """{jobId: colIdx} for the contiguous JOB header run starting at column B, same scan
    ai_data_write.py's own find_table uses."""
    cols = {}
    c = 2
    while ws.cell(row=header_row, column=c).value:
        cols[ws.cell(row=header_row, column=c).value] = c
        c += 1
    return cols


count_header_row = find_header_row('試行回数')
existing_cols = job_header_cols(count_header_row)
missing = [j for j in current_job_ids if j not in existing_cols]

if not missing:
    print('CONJOB sheet already has a column for every current JOB face -- nothing to do.')
    sys.exit(0)

last_col = max(existing_cols.values())
insert_at = last_col + 1
n = len(missing)
print(f'Inserting {n} new JOB column(s) at column {get_column_letter(insert_at)}: {missing}')

ws.insert_cols(insert_at, n)

avg_header_row = find_header_row('平均得点')
qst_avg_header_row = find_header_row('QST平均得点')
for header_row in (count_header_row, avg_header_row, qst_avg_header_row):
    for i, job_id in enumerate(missing):
        ws.cell(row=header_row, column=insert_at + i, value=job_id)

# Each hand-maintained aggregate row-triple sits 12 rows below its own table's header + 1 (data starts
# at header_row+1, runs for the 12 CON rows, then SUM/AVERAGE/label follow immediately) -- same "12 CON
# rows" both tables share (confirmed: CONJOB always has exactly 12 CON faces, one row each).
CON_ROW_COUNT = 12
for header_row in (avg_header_row, qst_avg_header_row):
    data_first_row = header_row + 1
    data_last_row = data_first_row + CON_ROW_COUNT - 1
    sum_row = data_last_row + 1
    avg_row = sum_row + 1
    label_row = avg_row + 1
    for i, job_id in enumerate(missing):
        col = insert_at + i
        col_letter = get_column_letter(col)
        ws.cell(row=sum_row, column=col, value=f'=SUM({col_letter}{data_first_row}:{col_letter}{data_last_row})')
        ws.cell(row=avg_row, column=col, value=f'={col_letter}{sum_row}/{CON_ROW_COUNT}')
        if job_id in PROPOSED_LABELS:
            ws.cell(row=label_row, column=col, value=PROPOSED_LABELS[job_id])

# 使用回数 row (single row, same JOB columns as count_header_row -- no per-JOB aggregate block of its own,
# just leave the new cells blank; the next real ai_data_write.py run fills them in like everything else).

wb.save(XLSX_PATH)
print(f'Saved {XLSX_PATH}. Run a fresh AI battle (ai_data_report.js + ai_data_write.py, or run_ai_battle.js)')
print('to fill in the new columns\' data and correct the per-CON marginal-average column\'s now-wider range.')
