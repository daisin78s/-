"""
One-time-per-addition structural migration for AI.DATA.xlsx's ABCM sheet, run by hand (2026-08-25, per
user request: "ABCカードはID順になるようにしてください") -- reorders every row into physical-ID order
(A/B/C decks each own tier-A block first, in game.json's own row order -- e.g. A001,A002,...,A006,A201,
A202,A301 -- then the same order again for tier-B, then M -- M001-M012 base monuments, then M401-403),
and relabels any row whose text no longer matches its card's current NAME (same drift
ai_data_sync_card_list.py's own append step can leave behind, e.g. when a card gets renamed after its
row was already added -- confirmed 2026-08-25: A202A/B006A/C301A's rows still said "訓練場の支配"/"移ろい
の兆し"/"王女" without the "LV1" suffix every other tier-A row already picked up when NAME started baking
that in). Preserves every row's own accumulated data (试行回数/平均得点/使用回数/etc, columns B onward) --
only the row's position and its own ID-column label move.

Unlike ai_data_sync_card_list.py (which only ever appends new rows at the end, run after every AI battle
as part of the regular pipeline), this changes the sheet's actual row ORDER, so it's meant to be run by
hand, once, whenever the row order has drifted out of ID order (new-card rows piling up at the bottom is
the main way that happens).

Usage: python tools/ai_data_reorder_abcm_rows.py [xlsxPath]
"""

import sys
import json
import openpyxl

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else 'AI.DATA.xlsx'
GAME_JSON_PATH = 'data/game.json'

with open(GAME_JSON_PATH, 'r', encoding='utf-8') as f:
    game_data = json.load(f)

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['ABCM']
max_col = ws.max_column

# {label text: [full row values, columns 1..max_col]} -- captured before anything is cleared, so every
# row's own accumulated data survives the move regardless of which label it's currently filed under.
existing_rows = {}
r = 2
while ws.cell(row=r, column=1).value:
    label = ws.cell(row=r, column=1).value
    existing_rows[label] = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
    r += 1
last_row = r - 1

# Target order: A/B/C decks' own tier-A faces in game.json's row order (already ID-ascending -- e.g.
# A001,A002,...,A006,A201,A202,A301), concatenated across the 3 decks; then the same for tier-B; then M
# (M001-M012, then M401-403, also already in file order). 'D' (not a real card, no NAME) is pinned first,
# matching its current position -- it was never part of the ID-ordering problem.
def names_for_tier(tier_suffix):
    names = []
    for sheet in ('A', 'B', 'C'):
        for row in game_data[sheet]:
            if row['ID'].endswith(tier_suffix):
                names.append(row['NAME'])
    return names

target_order = ['D'] + names_for_tier('A') + names_for_tier('B') + [row['NAME'] for row in game_data['M']]

# Every current row's label, resolved to its card's *current* NAME when it's stale (e.g. "訓練場の支配"
# -> "訓練場の支配LV1") by matching whichever target-order name it's a prefix/rename of via physicalId --
# simplest robust approach: build {physicalId: current NAME} from game.json, and {old label: physicalId}
# by finding which physicalId's OLD or NEW name equals the existing row's label.
name_by_id = {}
for sheet in ('A', 'B', 'C', 'M'):
    for row in game_data[sheet]:
        name_by_id[row['ID']] = row['NAME']

# Map every existing row's label to the physicalId it belongs to, so a stale label (old NAME) still
# resolves correctly -- tries an exact NAME match first (the common case), then falls back to prefix
# matching for a stale label that's missing a "LV1"/"LV2" suffix its current NAME has gained.
id_by_existing_label = {}
for label in existing_rows:
    match = None
    for id_, name in name_by_id.items():
        if name == label:
            match = id_
            break
    if match is None:
        for id_, name in name_by_id.items():
            if name.startswith(label) and name != label:
                match = id_
                break
    id_by_existing_label[label] = match

relabeled = []
unmatched = []
for label, id_ in id_by_existing_label.items():
    if label == 'D':
        continue  # not a real card, no physicalId/NAME to resolve -- see the write loop's own handling
    if id_ is None:
        unmatched.append(label)
    elif name_by_id[id_] != label:
        relabeled.append((label, name_by_id[id_]))

# Clear every existing row, then rewrite in target_order -- each row's own captured data follows its
# (possibly relabeled) name to its new position. A target name with no existing row (a card added since
# the last sync) is skipped here -- run ai_data_sync_card_list.py first if there's a genuinely new card,
# this script only reorders/relabels what's already present.
for r in range(2, last_row + 1):
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).value = None

written = 0
skipped_no_row = []
next_row = 2
for name in target_order:
    # 'D' (gained a new colored die) isn't a real card -- no physicalId/NAME to resolve, its existing
    # row (if any) carries straight over unchanged, matched by its own literal label.
    if name == 'D':
        src_label = 'D' if 'D' in existing_rows else None
    else:
        # Find the existing row whose (possibly stale) label resolves to this target name's physicalId.
        src_label = None
        for label, id_ in id_by_existing_label.items():
            if id_ is not None and name_by_id[id_] == name:
                src_label = label
                break
    if src_label is None:
        skipped_no_row.append(name)
        continue
    values = existing_rows[src_label]
    for c, v in enumerate(values, start=1):
        ws.cell(row=next_row, column=c).value = v if c > 1 else name  # column 1 gets the current NAME
    next_row += 1
    written += 1

print(f'Reordered {written} rows into ID order.')
if relabeled:
    print(f'Relabeled {len(relabeled)} stale row(s):')
    for old, new in relabeled:
        print(f'  {old!r} -> {new!r}')
if unmatched:
    print(f'WARNING: {len(unmatched)} existing row(s) had no matching card in game.json (left out): {unmatched}')
if skipped_no_row:
    print(f'WARNING: {len(skipped_no_row)} card(s) in game.json had no existing row to carry over (run ai_data_sync_card_list.py first for a genuinely new card): {skipped_no_row}')

wb.save(XLSX_PATH)
print(f'Saved {XLSX_PATH}.')
