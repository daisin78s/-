"""Convert game.xlsx (master data) into data/game.json (normalized runtime data) and
data/game.data.js (the same data, embedded as a browser global -- see below).

Usage: python tools/xlsx_to_json.py
Reads data/game.xlsx, writes data/game.json and data/game.data.js.
This script is a dev-time tool only; the game engine never reads xlsx directly.
"""
import json
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = PROJECT_ROOT / "data" / "game.xlsx"
DEST_JSON = PROJECT_ROOT / "data" / "game.json"
# Confirmed 2026-07-30: main.js runs from a file:// page (no build step, no local server -- the user
# wants to keep double-clicking index.html open), so it can't fs.readFileSync() or fetch() game.json
# the way Node (tests/tools) does. This file embeds the identical data as a plain browser global
# instead, loaded via a normal <script src="data/game.data.js"> tag -- see index.html and
# src/data-loader.js's loadGameData(), which accepts this pre-parsed object directly.
DEST_JS = PROJECT_ROOT / "data" / "game.data.js"

# MAP sheet's 3rd column has no header in the source spreadsheet.
UNNAMED_COLUMN_OVERRIDES = {
    ("MAP", 2): "CURRENT_AREA",
}


def load_sheet(ws):
    # The header is normally row 1, but a sheet can have a blank leading row (found 2026-08-27 in
    # 評価値_3, whose real header sits on row 2 -- silently produced all-empty rows under the old
    # header_row=row-1 assumption, since every column index came up unnamed and got skipped). Scan
    # forward to the first row that isn't entirely blank and use that as the header instead, so a
    # leading blank row never breaks a sheet's parse.
    header_row_num = 1
    for row in ws.iter_rows(min_row=1):
        if any(cell.value is not None for cell in row):
            header_row_num = row[0].row
            break

    header_row = ws[header_row_num]
    columns = []  # list of (index, name)
    for idx, cell in enumerate(header_row):
        name = cell.value
        if name is None:
            override = UNNAMED_COLUMN_OVERRIDES.get((ws.title, idx))
            if override is None:
                continue
            name = override
        columns.append((idx, name))

    rows = []
    for row in ws.iter_rows(min_row=header_row_num + 1):
        row_id = row[0].value
        if row_id is None:
            continue
        obj = {}
        for idx, name in columns:
            value = row[idx].value if idx < len(row) else None
            obj[name] = value if value is not None else ""
        rows.append(obj)
    return rows


def main():
    if not SRC_XLSX.exists():
        print(f"ERROR: source file not found: {SRC_XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    data = {}
    for ws in wb.worksheets:
        data[ws.title] = load_sheet(ws)

    DEST_JSON.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # `</script` inside any cell text would otherwise prematurely close the <script> tag this gets
    # loaded through -- escape it defensively even though nothing in the data currently contains it.
    inline_json = json.dumps(data, ensure_ascii=False).replace("</script", "<\\/script")
    DEST_JS.write_text(f"window.GAME_DATA = {inline_json};\n", encoding="utf-8")

    sheet_summary = ", ".join(f"{name}={len(rows)}" for name, rows in data.items())
    print(f"Wrote {DEST_JSON}")
    print(f"Wrote {DEST_JS}")
    print(sheet_summary)


if __name__ == "__main__":
    main()
