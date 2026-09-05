"""
Appends one generation's worth of GA training progress into a shared Excel workbook (tools/ga_train.js's
own per-generation checkpoint, 2026-09-04, per user request: "エクセルに 1 2 3 4とシートを作り（新しい
世代ほど先に来る） gameエクセルの勝率　平均点　（素点とQSTも） 最良個体の勝率　平均点　（素点とQSTも）
評価値を書きこむ"). Called once per generation from ga_train.js's own main loop (execFileSync), same
Node->Python subprocess pattern tools/ai_data_report.js already uses for tools/ai_data_write.py.

One sheet per generation, named by its own generation number, newest generation's sheet always moved to
the FIRST (leftmost) tab position. Each sheet has:
  - Row 1-2: "game.xlsx(アンカー)" -- its own 勝率/平均点(合計)/素点/QST, this generation's own fresh
    measurement (not a one-time baseline -- see ga_train.js's own ANCHOR_COUNT doc).
  - Row 3-4: "最良個体" -- same stats for this generation's own best individual (not necessarily the
    best-EVER across the whole run -- that's best_genome.json's own job).
  - Row 6+: the best individual's full 評価値-shaped table (ID, 1R, 2R, 3R, 4R), same layout
    tools/ga_genome_to_xlsx.py already uses for a single genome.

Usage:
    python tools/ga_progress_to_xlsx.py <outputDir> <generationSummaryJsonPath>

generationSummaryJsonPath is a small per-generation JSON (written fresh by ga_train.js each generation,
not the full gen_XXXX.json population dump) shaped:
    { "generation": int,
      "anchor": {"avgRank","avgScore","avgRawScore","avgQstScore","winRate"} | null,
      "best": {"avgRank","avgScore","avgRawScore","avgQstScore","winRate","genome"} }
"""

import json
import os
import sys
import openpyxl

if len(sys.argv) < 3:
    print('Usage: python tools/ga_progress_to_xlsx.py <outputDir> <generationSummaryJsonPath>')
    sys.exit(1)
OUTPUT_DIR = sys.argv[1]
SUMMARY_PATH = sys.argv[2]
XLSX_PATH = os.path.join(OUTPUT_DIR, 'progress.xlsx')

with open(SUMMARY_PATH, 'r', encoding='utf-8') as f:
    summary = json.load(f)

generation = summary['generation']
anchor = summary.get('anchor')
best = summary['best']
sheet_name = str(generation)

if os.path.exists(XLSX_PATH):
    wb = openpyxl.load_workbook(XLSX_PATH)
else:
    wb = openpyxl.Workbook()
    # openpyxl always creates one default "Sheet" on a brand-new Workbook() -- removed once we've added
    # our own first real sheet below, never left behind as a confusing empty extra tab.

if sheet_name in wb.sheetnames:
    del wb[sheet_name]  # re-running the same generation (e.g. a resumed/restarted run) replaces, not duplicates
ws = wb.create_sheet(sheet_name)

def write_stats_block(ws, start_row, label, stats):
    ws.cell(row=start_row, column=1, value=label)
    ws.cell(row=start_row + 1, column=1, value='勝率')
    ws.cell(row=start_row + 1, column=2, value=round(stats['winRate'], 3))
    ws.cell(row=start_row + 1, column=3, value='平均点(合計)')
    ws.cell(row=start_row + 1, column=4, value=round(stats['avgScore'], 2))
    ws.cell(row=start_row + 1, column=5, value='素点')
    ws.cell(row=start_row + 1, column=6, value=round(stats['avgRawScore'], 2))
    ws.cell(row=start_row + 1, column=7, value='QST')
    ws.cell(row=start_row + 1, column=8, value=round(stats['avgQstScore'], 2))
    ws.cell(row=start_row + 1, column=9, value='平均順位')
    ws.cell(row=start_row + 1, column=10, value=round(stats['avgRank'], 2))
    # 2026-09-04, per user request: "最良個体の平均点と平均順位 何試合での平均かも書いてほしい 例 100試合"
    # -- all 5 stats above (勝率/平均点/素点/QST/平均順位) come from the same gamesPlayed sample, so one
    # shared "試合数" column covers all of them rather than repeating it per metric.
    ws.cell(row=start_row + 1, column=11, value='試合数')
    ws.cell(row=start_row + 1, column=12, value=f"{stats['gamesPlayed']}試合")

row = 1
if anchor:
    write_stats_block(ws, row, 'game.xlsx(アンカー)', anchor)
    row += 2
write_stats_block(ws, row, '最良個体', best)
row += 3  # blank spacer row before the eval table

genome = best['genome']
ids = sorted(genome['1'].keys())
header_row = row
ws.cell(row=header_row, column=1, value='ID')
for i, r in enumerate((1, 2, 3, 4)):
    ws.cell(row=header_row, column=2 + i, value=f'{r}R')
row += 1
for id_ in ids:
    ws.cell(row=row, column=1, value=id_)
    for i, r in enumerate((1, 2, 3, 4)):
        value = genome[str(r)][id_]
        ws.cell(row=row, column=2 + i, value=round(value, 2) if isinstance(value, float) else value)
    row += 1

for col_letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'):
    ws.column_dimensions[col_letter].width = 14

# Newest generation always first (leftmost tab) -- move this sheet to index 0.
wb.move_sheet(sheet_name, offset=-wb.sheetnames.index(sheet_name))

# Remove openpyxl's default "Sheet" now that a real one exists (only present on a brand-new workbook,
# and only ever this exact default name/state -- never touches a real generation sheet).
if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1 and wb['Sheet'].max_column == 1 and wb['Sheet']['A1'].value is None:
    del wb['Sheet']

wb.save(XLSX_PATH)
