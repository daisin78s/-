"""
Exports a GA-trained genome (tools/ga_train.js's best_genome.json, or any gen_XXXX.json's own
per-individual genome field) into an .xlsx laid out exactly like game.xlsx's own 評価値 sheet
(ID column + 1R/2R/3R/4R columns) -- so it can be opened and eyeballed/compared in Excel the same way
the real hand-tuned table is, instead of only as raw JSON.

Usage:
    python tools/ga_genome_to_xlsx.py <genomeJsonPath> [outputXlsxPath]

genomeJsonPath may be either:
  - a best_genome.json (shape {generation, avgRank, avgScore, genome: {1:{...},2:{...},3:{...},4:{...}}})
  - a bare genome ({1:{...},2:{...},3:{...},4:{...}}, e.g. copy-pasted out of a gen_XXXX.json's own
    population[i].genome)

outputXlsxPath defaults to genomeJsonPath with its extension swapped to .xlsx.
"""

import json
import sys
import openpyxl

JSON_PATH = sys.argv[1] if len(sys.argv) > 1 else None
if not JSON_PATH:
    print('Usage: python tools/ga_genome_to_xlsx.py <genomeJsonPath> [outputXlsxPath]')
    sys.exit(1)
XLSX_PATH = sys.argv[2] if len(sys.argv) > 2 else JSON_PATH.rsplit('.', 1)[0] + '.xlsx'

with open(JSON_PATH, 'r', encoding='utf-8') as f:
    data = json.load(f)

# best_genome.json wraps the genome under its own "genome" key alongside generation/avgRank/avgScore;
# a bare genome (just {1:{...},...}) has no such wrapper -- detected by whether "genome" is present.
meta = None
if 'genome' in data:
    meta = {k: v for k, v in data.items() if k != 'genome'}
    genome = data['genome']
else:
    genome = data

ids = sorted(genome['1'].keys())

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '評価値'

row = 1
if meta:
    ws.cell(row=row, column=1, value='(metadata)')
    row += 1
    for key, value in meta.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1  # blank spacer row before the real table

header_row = row
ws.cell(row=header_row, column=1, value='ID')
for i, r in enumerate((1, 2, 3, 4)):
    ws.cell(row=header_row, column=2 + i, value=f'{r}R')
row += 1

for id_ in ids:
    ws.cell(row=row, column=1, value=id_)
    for i, r in enumerate((1, 2, 3, 4)):
        value = genome[str(r)][id_]
        # Rounded for readability (2026-09-04: raw genome values carry long float tails from
        # mutateGenomePercent's proportional nudging, e.g. 31.588775767944753) -- 2 decimal places is
        # plenty of precision for eyeballing/comparing against the real sheet's own whole-number values.
        ws.cell(row=row, column=2 + i, value=round(value, 2) if isinstance(value, float) else value)
    row += 1

for col_letter in ('A', 'B', 'C', 'D', 'E'):
    ws.column_dimensions[col_letter].width = 16

wb.save(XLSX_PATH)
print(f'Wrote {XLSX_PATH} ({len(ids)} ids x 4 rounds)' + (f', generation={meta.get("generation")} avgRank={meta.get("avgRank")} avgScore={meta.get("avgScore")}' if meta else ''))
